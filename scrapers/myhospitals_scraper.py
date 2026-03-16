#!/usr/bin/env python3
"""
MyHospitals Scraper — Enrich hospitals with verified bed counts and types.

Source: AIHW MyHospitals dataset (published CSV/Excel files)
  - https://www.aihw.gov.au/reports-data/myhospitals
  - Direct data downloads from AIHW data portal

Matching: Fuzzy name + state matching against our hospitals table.
Updates: bed_count, hospital_type, bed_count_verified
"""

import argparse
import csv
import io
import json
import logging
import os
import re
import sqlite3
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DB_PATH = Path(__file__).resolve().parent.parent / "pharmacy_finder.db"
CACHE_DIR = Path(__file__).resolve().parent / "cache" / "myhospitals"

# AIHW data URLs (these are the public data download pages)
AIHW_MYHOSPITALS_URL = "https://www.aihw.gov.au/reports-data/myhospitals"
AIHW_DATA_URL = "https://www.aihw.gov.au/reports-data/myhospitals/sectors/overview"
AIHW_HOSPITAL_RESOURCES_URL = "https://www.aihw.gov.au/reports-data/myhospitals/intersection/activity/apc"

# Known direct data download URLs (AIHW publishes these periodically)
# These may change — the scraper will try to find current ones
AIHW_DATA_DOWNLOADS = [
    "https://www.aihw.gov.au/getmedia/myhospitals-data.csv",
    "https://www.aihw.gov.au/reports-data/myhospitals/content/data-downloads",
]

# Alternative: MyHospitals API (if available)
MYHOSPITALS_API = "https://www.myhospitals.gov.au/api/data"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

log = logging.getLogger("myhospitals_scraper")

# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def ensure_columns(conn: sqlite3.Connection):
    """Add enrichment columns if they don't exist."""
    cursor = conn.cursor()
    existing = {row[1] for row in cursor.execute("PRAGMA table_info(hospitals)").fetchall()}
    if "bed_count_verified" not in existing:
        cursor.execute("ALTER TABLE hospitals ADD COLUMN bed_count_verified INTEGER DEFAULT 0")
        log.info("Added column bed_count_verified")
    conn.commit()


def get_hospitals(conn: sqlite3.Connection, state: str = None, limit: int = None):
    """Fetch hospitals to enrich."""
    sql = "SELECT id, name, address, bed_count, hospital_type FROM hospitals WHERE 1=1"
    params = []
    if state:
        sql += " AND address LIKE ?"
        params.append(f"% {state.upper()} %")
    sql += " ORDER BY id"
    if limit:
        sql += " LIMIT ?"
        params.append(limit)
    return conn.execute(sql, params).fetchall()


def update_hospital(conn: sqlite3.Connection, hospital_id: int, bed_count: int,
                    hospital_type: str, source: str, dry_run: bool = False):
    """Update a hospital with verified data."""
    if dry_run:
        log.info(f"[DRY-RUN] Would update id={hospital_id}: beds={bed_count}, type={hospital_type}")
        return
    conn.execute("""
        UPDATE hospitals
        SET bed_count = ?, hospital_type = ?, bed_count_verified = 1
        WHERE id = ?
    """, (bed_count, hospital_type, hospital_id))
    conn.commit()


# ---------------------------------------------------------------------------
# Name matching
# ---------------------------------------------------------------------------

def _normalise_hospital_name(name: str) -> str:
    """Normalise hospital name for matching."""
    name = name.lower().strip()
    # Remove common suffixes
    for term in ["hospital", "health service", "health campus", "campus",
                 "medical centre", "private", "public", "base"]:
        name = name.replace(term, "")
    # Remove state abbreviations
    name = re.sub(r"\b(nsw|vic|qld|sa|wa|tas|nt|act)\b", "", name)
    return re.sub(r"\s+", " ", name).strip()


def _name_similarity(a: str, b: str) -> float:
    """Word-overlap similarity 0-1."""
    wa = set(_normalise_hospital_name(a).split())
    wb = set(_normalise_hospital_name(b).split())
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / max(len(wa), len(wb))


def _extract_state(address: str) -> str:
    """Extract state abbreviation from address."""
    m = re.search(r"\b(NSW|VIC|QLD|SA|WA|TAS|NT|ACT)\b", address)
    return m.group(1) if m else ""


# ---------------------------------------------------------------------------
# Data acquisition
# ---------------------------------------------------------------------------

def download_aihw_data(session: requests.Session) -> list[dict]:
    """Attempt to download AIHW MyHospitals dataset.
    
    Tries multiple approaches:
    1. Direct CSV/Excel download links
    2. Scrape the data downloads page for current links
    3. Scrape individual hospital pages from myhospitals.gov.au
    
    Returns list of dicts with keys: name, state, bed_count, hospital_type, sector
    """
    hospitals_data = []
    
    # Approach 1: Try myhospitals.gov.au website
    hospitals_data = _scrape_myhospitals_website(session)
    if hospitals_data:
        return hospitals_data
    
    # Approach 2: Try AIHW data downloads page
    hospitals_data = _find_aihw_downloads(session)
    if hospitals_data:
        return hospitals_data
    
    # Approach 3: Try direct known download URLs
    for url in AIHW_DATA_DOWNLOADS:
        try:
            resp = session.get(url, headers=HEADERS, timeout=30)
            if resp.status_code == 200:
                content_type = resp.headers.get("Content-Type", "")
                if "csv" in content_type or url.endswith(".csv"):
                    hospitals_data = _parse_csv_data(resp.text)
                    if hospitals_data:
                        return hospitals_data
        except requests.RequestException:
            continue
    
    return hospitals_data


def _scrape_myhospitals_website(session: requests.Session) -> list[dict]:
    """Scrape myhospitals.gov.au for hospital data."""
    hospitals = []
    
    try:
        # Try the MyHospitals website
        resp = session.get("https://www.myhospitals.gov.au/hospital/search",
                          headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            # Try alternative URL
            resp = session.get("https://www.myhospitals.gov.au/about-the-data/download-data",
                              headers=HEADERS, timeout=15)
        
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "lxml")
            
            # Look for download links (CSV, Excel)
            for a in soup.select("a[href]"):
                href = a.get("href", "")
                if any(ext in href.lower() for ext in [".csv", ".xlsx", ".xls"]):
                    full_url = urljoin("https://www.myhospitals.gov.au", href)
                    log.info(f"Found data download: {full_url}")
                    data = _download_and_parse(full_url, session)
                    if data:
                        return data
            
            # Look for hospital listings on the page
            for el in soup.select(".hospital-item, .result-item, tr"):
                name_el = el.select_one(".hospital-name, .name, td:first-child")
                if name_el:
                    name = name_el.get_text(strip=True)
                    if len(name) > 3:
                        # Try to extract beds and type
                        text = el.get_text()
                        beds_match = re.search(r"(\d+)\s*beds?", text, re.I)
                        type_match = re.search(r"\b(public|private)\b", text, re.I)
                        state_match = re.search(r"\b(NSW|VIC|QLD|SA|WA|TAS|NT|ACT)\b", text)
                        
                        hospitals.append({
                            "name": name,
                            "state": state_match.group(1) if state_match else "",
                            "bed_count": int(beds_match.group(1)) if beds_match else 0,
                            "hospital_type": type_match.group(1).lower() if type_match else "",
                            "sector": type_match.group(1).lower() if type_match else "",
                        })
    except requests.RequestException as e:
        log.debug(f"MyHospitals website scrape failed: {e}")
    
    return hospitals


def _find_aihw_downloads(session: requests.Session) -> list[dict]:
    """Scrape AIHW data page for download links."""
    try:
        resp = session.get(AIHW_MYHOSPITALS_URL, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            return []
        
        soup = BeautifulSoup(resp.text, "lxml")
        
        for a in soup.select("a[href]"):
            href = a.get("href", "")
            text = a.get_text(strip=True).lower()
            if any(kw in text for kw in ["download", "data", "hospital"]):
                if any(ext in href.lower() for ext in [".csv", ".xlsx", ".xls", "download"]):
                    full_url = urljoin(AIHW_MYHOSPITALS_URL, href)
                    log.info(f"Found potential AIHW download: {full_url}")
                    data = _download_and_parse(full_url, session)
                    if data:
                        return data
    except requests.RequestException as e:
        log.debug(f"AIHW data page scrape failed: {e}")
    
    return []


def _download_and_parse(url: str, session: requests.Session) -> list[dict]:
    """Download a file and parse it for hospital data."""
    try:
        resp = session.get(url, headers=HEADERS, timeout=30)
        if resp.status_code != 200:
            return []
        
        # Try CSV parsing
        if url.endswith(".csv") or "csv" in resp.headers.get("Content-Type", ""):
            return _parse_csv_data(resp.text)
        
        # Try Excel parsing (requires openpyxl)
        if url.endswith(".xlsx") or url.endswith(".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(resp.content))
                ws = wb.active
                return _parse_excel_data(ws)
            except ImportError:
                log.warning("openpyxl not installed — can't parse Excel files")
            except Exception as e:
                log.debug(f"Excel parsing failed: {e}")
    except requests.RequestException:
        pass
    
    return []


def _parse_csv_data(text: str) -> list[dict]:
    """Parse CSV hospital data."""
    hospitals = []
    reader = csv.DictReader(io.StringIO(text))
    
    for row in reader:
        # Try common column name patterns
        name = (row.get("Hospital Name") or row.get("hospital_name") or
                row.get("Name") or row.get("name") or
                row.get("Establishment name") or "")
        state = (row.get("State") or row.get("state") or
                 row.get("State/Territory") or row.get("Jurisdiction") or "")
        beds = (row.get("Beds") or row.get("beds") or
                row.get("Available beds") or row.get("Bed count") or
                row.get("Number of beds") or "0")
        sector = (row.get("Sector") or row.get("sector") or
                  row.get("Hospital type") or row.get("Type") or
                  row.get("Ownership") or "")
        
        if name and len(name) > 3:
            try:
                bed_count = int(re.sub(r"[^\d]", "", str(beds)) or "0")
            except ValueError:
                bed_count = 0
            
            hospitals.append({
                "name": name.strip(),
                "state": state.strip().upper(),
                "bed_count": bed_count,
                "hospital_type": _normalise_type(sector),
                "sector": sector.strip(),
            })
    
    return hospitals


def _parse_excel_data(ws) -> list[dict]:
    """Parse Excel worksheet for hospital data."""
    hospitals = []
    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
    
    # Find relevant columns
    name_col = next((i for i, h in enumerate(headers)
                     if any(kw in h for kw in ["hospital", "name", "establishment"])), None)
    state_col = next((i for i, h in enumerate(headers)
                      if any(kw in h for kw in ["state", "jurisdiction", "territory"])), None)
    beds_col = next((i for i, h in enumerate(headers)
                     if any(kw in h for kw in ["bed", "beds"])), None)
    type_col = next((i for i, h in enumerate(headers)
                     if any(kw in h for kw in ["sector", "type", "ownership"])), None)
    
    if name_col is None:
        return []
    
    for row_idx in range(2, ws.max_row + 1):
        cells = [ws.cell(row=row_idx, column=i + 1).value for i in range(len(headers))]
        name = str(cells[name_col] or "").strip()
        if not name or len(name) < 3:
            continue
        
        hospitals.append({
            "name": name,
            "state": str(cells[state_col] or "").strip().upper() if state_col else "",
            "bed_count": int(cells[beds_col] or 0) if beds_col and cells[beds_col] else 0,
            "hospital_type": _normalise_type(str(cells[type_col] or "")) if type_col else "",
            "sector": str(cells[type_col] or "").strip() if type_col else "",
        })
    
    return hospitals


def _normalise_type(sector: str) -> str:
    """Normalise hospital type to 'public' or 'private'."""
    sector = sector.lower().strip()
    if "public" in sector or "government" in sector:
        return "public"
    if "private" in sector or "not-for-profit" in sector or "nfp" in sector:
        return "private"
    return sector


# ---------------------------------------------------------------------------
# Fallback: scrape individual hospital pages
# ---------------------------------------------------------------------------

def scrape_individual_hospital(name: str, state: str, session: requests.Session) -> dict | None:
    """Scrape data for a single hospital from myhospitals.gov.au or AIHW."""
    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    
    urls_to_try = [
        f"https://www.myhospitals.gov.au/hospital/{slug}",
        f"https://www.myhospitals.gov.au/hospital/{slug}/overview",
    ]
    
    for url in urls_to_try:
        try:
            resp = session.get(url, headers=HEADERS, timeout=15, allow_redirects=True)
            if resp.status_code != 200:
                continue
            
            soup = BeautifulSoup(resp.text, "lxml")
            text = soup.get_text()
            
            beds_match = re.search(r"(\d+)\s*(?:available\s+)?beds?", text, re.I)
            type_match = re.search(r"\b(public|private|not.for.profit)\b", text, re.I)
            
            if beds_match:
                return {
                    "name": name,
                    "bed_count": int(beds_match.group(1)),
                    "hospital_type": _normalise_type(type_match.group(1)) if type_match else "",
                    "source_url": url,
                }
        except requests.RequestException:
            continue
    
    return None


# ---------------------------------------------------------------------------
# Cache
# ---------------------------------------------------------------------------

def _cache_path(hospital_id: int) -> Path:
    return CACHE_DIR / f"hospital_{hospital_id}.json"


def _bulk_cache_path() -> Path:
    return CACHE_DIR / "aihw_bulk_data.json"


def save_cache(hospital_id: int, data: dict):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    with open(_cache_path(hospital_id), "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def load_cache(hospital_id: int) -> dict | None:
    p = _cache_path(hospital_id)
    if p.exists():
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def save_bulk_cache(data: list[dict]):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    with open(_bulk_cache_path(), "w", encoding="utf-8") as f:
        json.dump({"downloaded_at": datetime.now().isoformat(), "hospitals": data}, f, indent=2)


def load_bulk_cache() -> list[dict] | None:
    p = _bulk_cache_path()
    if p.exists():
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
            # Cache valid for 30 days
            downloaded = data.get("downloaded_at", "")
            if downloaded:
                try:
                    dt = datetime.fromisoformat(downloaded)
                    if (datetime.now() - dt).days < 30:
                        return data.get("hospitals", [])
                except ValueError:
                    pass
    return None


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

def match_hospitals(our_hospitals: list, aihw_data: list[dict]) -> dict:
    """Match our hospital records against AIHW data.
    
    Returns dict mapping our hospital_id -> aihw record.
    """
    matches = {}
    
    for row in our_hospitals:
        hospital_id, name, address, bed_count, hospital_type = row
        our_state = _extract_state(address)
        
        best_match = None
        best_score = 0.0
        
        for aihw in aihw_data:
            # State must match (if available)
            if our_state and aihw.get("state") and our_state != aihw["state"]:
                continue
            
            score = _name_similarity(name, aihw["name"])
            if score > best_score and score >= 0.4:
                best_score = score
                best_match = aihw
        
        if best_match:
            matches[hospital_id] = {
                **best_match,
                "match_score": best_score,
            }
            log.debug(f"Matched '{name}' -> '{best_match['name']}' (score={best_score:.2f})")
        else:
            log.debug(f"No match for '{name}'")
    
    return matches


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(state: str = None, limit: int = None, dry_run: bool = False):
    """Main entry point."""
    conn = sqlite3.connect(str(DB_PATH))
    ensure_columns(conn)

    our_hospitals = get_hospitals(conn, state=state, limit=limit)
    log.info(f"Processing {len(our_hospitals)} hospitals")

    session = requests.Session()
    stats = {"updated": 0, "failed": 0, "skipped": 0, "from_bulk": 0, "from_individual": 0}

    # Step 1: Try bulk AIHW data download
    log.info("Attempting bulk AIHW data download...")
    aihw_data = load_bulk_cache()
    if aihw_data:
        log.info(f"Using cached AIHW data ({len(aihw_data)} hospitals)")
    else:
        aihw_data = download_aihw_data(session)
        if aihw_data:
            save_bulk_cache(aihw_data)
            log.info(f"Downloaded {len(aihw_data)} hospitals from AIHW")
        else:
            log.warning("Could not download bulk AIHW data — will try individual lookups")
            aihw_data = []

    # Step 2: Match and update
    if aihw_data:
        matches = match_hospitals(our_hospitals, aihw_data)
        log.info(f"Matched {len(matches)} / {len(our_hospitals)} hospitals from bulk data")
        
        for row in our_hospitals:
            hospital_id, name, address, current_beds, current_type = row
            
            if hospital_id in matches:
                match = matches[hospital_id]
                bed_count = match.get("bed_count", 0) or current_beds or 0
                h_type = match.get("hospital_type", "") or current_type or ""
                
                if bed_count > 0:
                    save_cache(hospital_id, match)
                    update_hospital(conn, hospital_id, bed_count, h_type,
                                   source="aihw_bulk", dry_run=dry_run)
                    stats["updated"] += 1
                    stats["from_bulk"] += 1
                    log.info(f"[{hospital_id}] {name}: {bed_count} beds ({h_type}) — from bulk")
                else:
                    stats["skipped"] += 1
            else:
                # Try individual scrape
                cached = load_cache(hospital_id)
                if cached and cached.get("bed_count", 0) > 0:
                    update_hospital(conn, hospital_id, cached["bed_count"],
                                   cached.get("hospital_type", current_type or ""),
                                   source="cache", dry_run=dry_run)
                    stats["updated"] += 1
                    continue
                
                log.info(f"[{hospital_id}] {name}: trying individual lookup...")
                time.sleep(1)
                result = scrape_individual_hospital(name, _extract_state(address), session)
                if result and result.get("bed_count", 0) > 0:
                    save_cache(hospital_id, result)
                    update_hospital(conn, hospital_id, result["bed_count"],
                                   result.get("hospital_type", current_type or ""),
                                   source=f"myhospitals:{result.get('source_url', '')}",
                                   dry_run=dry_run)
                    stats["updated"] += 1
                    stats["from_individual"] += 1
                    log.info(f"  Found: {result['bed_count']} beds")
                else:
                    stats["skipped"] += 1
                    log.info(f"  No data found")
    else:
        # No bulk data — try individual lookups for each
        for row in our_hospitals:
            hospital_id, name, address, current_beds, current_type = row
            
            cached = load_cache(hospital_id)
            if cached and cached.get("bed_count", 0) > 0:
                update_hospital(conn, hospital_id, cached["bed_count"],
                               cached.get("hospital_type", current_type or ""),
                               source="cache", dry_run=dry_run)
                stats["updated"] += 1
                continue
            
            log.info(f"[{hospital_id}] {name}: individual lookup...")
            time.sleep(1)
            result = scrape_individual_hospital(name, _extract_state(address), session)
            if result and result.get("bed_count", 0) > 0:
                save_cache(hospital_id, result)
                update_hospital(conn, hospital_id, result["bed_count"],
                               result.get("hospital_type", current_type or ""),
                               source=f"myhospitals:{result.get('source_url', '')}",
                               dry_run=dry_run)
                stats["updated"] += 1
                stats["from_individual"] += 1
            else:
                stats["skipped"] += 1

    conn.close()

    # Summary
    log.info("=" * 50)
    log.info("MYHOSPITALS SCRAPER — SUMMARY")
    log.info(f"  Total processed: {len(our_hospitals)}")
    log.info(f"  Updated:         {stats['updated']}")
    log.info(f"    From bulk:     {stats['from_bulk']}")
    log.info(f"    Individual:    {stats['from_individual']}")
    log.info(f"  Skipped:         {stats['skipped']}")
    log.info(f"  Failed:          {stats['failed']}")
    log.info("=" * 50)

    return stats


def main():
    parser = argparse.ArgumentParser(description="Download AIHW MyHospitals data for bed counts")
    parser.add_argument("--state", type=str, help="Filter by state (e.g. TAS, VIC)")
    parser.add_argument("--limit", type=int, help="Max hospitals to process")
    parser.add_argument("--dry-run", action="store_true", help="Don't write to DB")
    parser.add_argument("--verbose", "-v", action="store_true", help="Debug logging")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )

    run(state=args.state, limit=args.limit, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
